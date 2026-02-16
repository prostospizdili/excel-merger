#!/usr/bin/env python3
"""
Excel Data Merger - Десктопное приложение для создания сводных таблиц
Автор: AI Assistant
Версия: 1.0

Инструкция по запуску:
1. Установите Python 3.9+ с python.org
2. Установите зависимости: pip install openpyxl customtkinter
3. Запустите: python excel_merger_app.py

Для создания .exe:
pip install pyinstaller
pyinstaller --onefile --windowed --name "ExcelMerger" excel_merger_app.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import customtkinter as ctk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json
import os
import threading
from typing import Dict, List, Set, Optional, Tuple
from pathlib import Path

# ==================== КОНСТАНТЫ ====================
CONFIG_FILE = "excel_merger_config.json"

# Стили
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


# ==================== МОДЕЛИ ДАННЫХ ====================
class SourceFile:
    """Модель исходного файла"""
    def __init__(self, file_path: str):
        self.id = f"file-{id(self)}"
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.display_name = os.path.splitext(self.file_name)[0]
        self.sheet_name = ""
        self.available_sheets: List[str] = []
        self.headers: List[str] = []
        self.preview: List[List[str]] = []
        self.mapping = {
            "vendorColumn": "",
            "statusColumn": "",
            "partNumberColumn": "",
            "dataColumn": ""
        }
        self._load_sheets()

    def _load_sheets(self):
        """Загружает список листов и превью данных"""
        try:
            # Используем read_only для больших файлов
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            self.available_sheets = wb.sheetnames.copy()
            self.sheet_name = self.available_sheets[0] if self.available_sheets else ""
            wb.close()

            # Загружаем превью отдельно
            self._load_preview()
        except Exception as e:
            raise Exception(f"Ошибка при чтении файла: {e}")

    def _load_preview(self):
        """Загружает превью данных (первые 20 строк)"""
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb[self.sheet_name]
            
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 20:
                    break
                rows.append([str(cell) if cell is not None else "" for cell in row])
            
            if rows:
                self.headers = rows[0] if rows else []
                self.preview = rows
            
            wb.close()
        except Exception as e:
            print(f"Ошибка при загрузке превью: {e}")

    def update_sheet(self, sheet_name: str):
        """Обновляет лист и перезагружает превью"""
        self.sheet_name = sheet_name
        self._load_preview()

    def to_dict(self) -> dict:
        return {
            "file_path": self.file_path,
            "display_name": self.display_name,
            "sheet_name": self.sheet_name,
            "mapping": self.mapping
        }

    @staticmethod
    def from_dict(data: dict) -> 'SourceFile':
        """Восстанавливает файл из словаря (без загрузки данных)"""
        if not os.path.exists(data.get("file_path", "")):
            return None
        try:
            sf = SourceFile(data["file_path"])
            sf.display_name = data.get("display_name", sf.display_name)
            if data.get("sheet_name") in sf.available_sheets:
                sf.update_sheet(data["sheet_name"])
            sf.mapping = data.get("mapping", sf.mapping)
            return sf
        except:
            return None


class ColumnFilter:
    """Модель фильтра столбца"""
    def __init__(self, source_file_id: str, file_display_name: str):
        self.id = f"filter-{id(self)}"
        self.source_file_id = source_file_id
        self.vendor_name = ""
        self.status_value = "1"
        self.column_name = f"Вендор_{file_display_name}"
        self.extract_data = False

    def to_dict(self) -> dict:
        return {
            "source_file_id": self.source_file_id,
            "vendor_name": self.vendor_name,
            "status_value": self.status_value,
            "column_name": self.column_name,
            "extract_data": self.extract_data
        }

    @staticmethod
    def from_dict(data: dict) -> 'ColumnFilter':
        cf = ColumnFilter("", "")
        cf.id = data.get("id", cf.id)
        cf.source_file_id = data.get("source_file_id", "")
        cf.vendor_name = data.get("vendor_name", "")
        cf.status_value = data.get("status_value", "1")
        cf.column_name = data.get("column_name", "")
        cf.extract_data = data.get("extract_data", False)
        return cf


class Counter:
    """Модель счётчика"""
    def __init__(self):
        self.id = f"counter-{id(self)}"
        self.name = "Итого"
        self.type = "SUM"
        self.target_column = ""
        self.source_columns: List[str] = []

    def to_dict(self) -> dict:
        return {
            "name": self.name,
            "type": self.type,
            "target_column": self.target_column,
            "source_columns": self.source_columns
        }

    @staticmethod
    def from_dict(data: dict) -> 'Counter':
        c = Counter()
        c.name = data.get("name", "Итого")
        c.type = data.get("type", "SUM")
        c.target_column = data.get("target_column", "")
        c.source_columns = data.get("source_columns", [])
        return c


# ==================== ОБРАБОТЧИК ДАННЫХ ====================
class DataProcessor:
    """Обработчик данных - работает с большими файлами"""

    @staticmethod
    def get_part_prefix(part_number: str) -> Optional[str]:
        """Извлекает префикс номера детали (до _)"""
        if not part_number:
            return None
        part_str = str(part_number).strip()
        if '_' in part_str:
            return part_str.split('_')[0].upper()
        return part_str.upper()

    @staticmethod
    def get_warehouse_type(part_number: str, row_labels: List[str]) -> Optional[str]:
        """Определяет тип склада по префиксу детали"""
        if not part_number:
            return None
        part_str = str(part_number).strip().upper()
        
        for label in row_labels:
            if part_str.startswith(label.upper()):
                return label
        return None

    @staticmethod
    def column_letter_to_index(letter: str) -> int:
        """Преобразует букву столбца в индекс (A=1, B=2, ...)"""
        result = 0
        for char in letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    @staticmethod
    def process_file(file: SourceFile, row_labels: List[str], 
                     progress_callback=None) -> Dict[str, Dict[str, Set[str]]]:
        """
        Обрабатывает файл в режиме read_only для экономии памяти
        Возвращает: {(vendor, status): {warehouse_type: set(part_prefixes)}}
        """
        result = defaultdict(lambda: defaultdict(set))
        
        if not file.mapping.get("vendorColumn") or not file.mapping.get("statusColumn") or not file.mapping.get("partNumberColumn"):
            return dict(result)

        vendor_col = DataProcessor.column_letter_to_index(file.mapping["vendorColumn"])
        status_col = DataProcessor.column_letter_to_index(file.mapping["statusColumn"])
        part_col = DataProcessor.column_letter_to_index(file.mapping["partNumberColumn"])

        try:
            # Открываем в режиме read_only для больших файлов
            wb = load_workbook(file.file_path, read_only=True, data_only=True)
            ws = wb[file.sheet_name]

            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
                row_count += 1
                
                # Прогресс каждые 1000 строк
                if progress_callback and row_count % 1000 == 0:
                    progress_callback(row_count)

                try:
                    vendor = str(row[vendor_col - 1] or "").strip()
                    status = str(row[status_col - 1] or "").strip()
                    part_number = str(row[part_col - 1] or "").strip()

                    if not vendor or not status or not part_number:
                        continue

                    # Определяем тип склада
                    wh_type = DataProcessor.get_warehouse_type(part_number, row_labels)
                    if wh_type:
                        prefix = DataProcessor.get_part_prefix(part_number)
                        if prefix:
                            key = (vendor, status)
                            result[key][wh_type].add(prefix)
                except IndexError:
                    continue

            wb.close()
            
        except Exception as e:
            raise Exception(f"Ошибка при обработке файла {file.file_name}: {e}")

        return dict(result)

    @staticmethod
    def create_summary_excel(row_labels: List[str], column_filters: List[ColumnFilter],
                             counters: List[Counter], processed_data: Dict[str, dict],
                             output_path: str, progress_callback=None) -> str:
        """Создаёт сводный Excel-файл"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводная таблица"

        # Заголовки
        ws.cell(row=1, column=1, value="Тип склада")
        for idx, col_filter in enumerate(column_filters, start=2):
            ws.cell(row=1, column=idx, value=col_filter.column_name)

        # Данные
        for row_idx, row_label in enumerate(row_labels, start=2):
            ws.cell(row=row_idx, column=1, value=row_label)
            
            for col_idx, col_filter in enumerate(column_filters, start=2):
                file_data = processed_data.get(col_filter.source_file_id, {})
                key = (col_filter.vendor_name, col_filter.status_value)
                
                if key in file_data and row_label in file_data[key]:
                    count = len(file_data[key][row_label])
                    ws.cell(row=row_idx, column=col_idx, value=count)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=0)

        # Строка итогов
        data_end_row = len(row_labels) + 1
        total_row = data_end_row + 1
        ws.cell(row=total_row, column=1, value="ИТОГО")

        for col_idx in range(2, len(column_filters) + 2):
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}2:{col_letter}{data_end_row})"
            ws.cell(row=total_row, column=col_idx, value=formula)

        # Форматирование
        DataProcessor._apply_formatting(ws, total_row, len(column_filters) + 1)

        wb.save(output_path)
        return output_path

    @staticmethod
    def _apply_formatting(ws, data_rows: int, data_cols: int):
        """Применяет форматирование к таблице"""
        # Заголовок
        for col in range(1, data_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        # Данные
        for row in range(2, data_rows + 1):
            for col in range(1, data_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = BORDER
                if col == 1:
                    cell.font = Font(bold=True)
                cell.alignment = CENTER_ALIGN

        # Автоширина
        for col in range(1, data_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max(max_length + 2, 12)

        # Автофильтр
        ws.auto_filter.ref = f"A1:{get_column_letter(data_cols)}{data_rows}"


# ==================== ГЛАВНОЕ ОКНО ПРИЛОЖЕНИЯ ====================
class ExcelMergerApp(ctk.CTk):
    """Главное окно приложения"""

    def __init__(self):
        super().__init__()

        # Настройка окна
        self.title("📊 Excel Data Merger")
        self.geometry("1200x900")
        self.minsize(900, 700)

        # Тема
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")

        # Данные
        self.row_labels: List[str] = ["LL", "LM", "LD", "LB"]
        self.source_files: List[SourceFile] = []
        self.column_filters: List[ColumnFilter] = []
        self.counters: List[Counter] = []

        # Создаём интерфейс
        self._create_ui()

        # Загружаем сохранённые настройки
        self._load_config()

    def _create_ui(self):
        """Создаёт интерфейс"""
        # Главный контейнер с прокруткой
        self.main_frame = ctk.CTkScrollableFrame(self, label_text="Создание сводных таблиц из Excel-файлов")
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # === СЕКЦИЯ 1: Строки ===
        self._create_row_labels_section()

        # === СЕКЦИЯ 2: Файлы ===
        self._create_files_section()

        # === СЕКЦИЯ 3: Столбцы ===
        self._create_filters_section()

        # === СЕКЦИЯ 4: Генерация ===
        self._create_generation_section()

    def _create_row_labels_section(self):
        """Секция настройки строк"""
        frame = ctk.CTkFrame(self.main_frame)
        frame.pack(fill="x", pady=(0, 10))

        # Заголовок
        header = ctk.CTkFrame(frame, fg_color="transparent")
        header.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(header, text="1. Строки сводной таблицы", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        ctk.CTkLabel(header, text="Типы складов (например: LL, LM, LD, LB)", 
                    text_color="gray").pack(side="left", padx=10)

        # Контейнер для строк
        self.row_labels_frame = ctk.CTkFrame(frame, fg_color="transparent")
        self.row_labels_frame.pack(fill="x", padx=10, pady=(0, 10))

        # Кнопка добавления
        ctk.CTkButton(self.row_labels_frame, text="+ Добавить строку", 
                     command=self._add_row_label, width=150).pack(side="left", padx=5)

        self._render_row_labels()

    def _create_files_section(self):
        """Секция загрузки файлов"""
        frame = ctk.CTkFrame(self.main_frame)
        frame.pack(fill="x", pady=(0, 10))

        # Заголовок
        header = ctk.CTkFrame(frame, fg_color="transparent")
        header.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(header, text="2. Исходные файлы", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        ctk.CTkLabel(header, text="Загрузите Excel-файлы для обработки", 
                    text_color="gray").pack(side="left", padx=10)

        # Кнопка загрузки
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkButton(btn_frame, text="📁 Выбрать файлы", 
                     command=self._load_files, width=150).pack(side="left")

        # Контейнер для файлов
        self.files_frame = ctk.CTkFrame(frame, fg_color="transparent")
        self.files_frame.pack(fill="x", padx=10, pady=(0, 10))

    def _create_filters_section(self):
        """Секция настройки столбцов"""
        frame = ctk.CTkFrame(self.main_frame)
        frame.pack(fill="x", pady=(0, 10))

        # Заголовок
        header = ctk.CTkFrame(frame, fg_color="transparent")
        header.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(header, text="3. Столбцы сводной таблицы", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        ctk.CTkLabel(header, text="Настройте фильтры (Вендор + Файл)", 
                    text_color="gray").pack(side="left", padx=10)

        # Кнопка добавления
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=5)
        
        self.add_filter_btn = ctk.CTkButton(btn_frame, text="+ Добавить столбец", 
                                           command=self._add_column_filter, width=150)
        self.add_filter_btn.pack(side="left")

        # Контейнер для фильтров
        self.filters_frame = ctk.CTkFrame(frame, fg_color="transparent")
        self.filters_frame.pack(fill="x", padx=10, pady=(0, 10))

    def _create_generation_section(self):
        """Секция генерации"""
        frame = ctk.CTkFrame(self.main_frame, fg_color=("gray90", "gray20"))
        frame.pack(fill="x", pady=10)

        # Кнопка генерации
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.pack(pady=20)
        
        self.generate_btn = ctk.CTkButton(btn_frame, text="📊 Сгенерировать сводный Excel", 
                                         command=self._generate_excel, width=250, height=40,
                                         font=ctk.CTkFont(size=14, weight="bold"))
        self.generate_btn.pack()

        # Прогресс
        self.progress_bar = ctk.CTkProgressBar(btn_frame, width=400)
        self.progress_bar.set(0)
        self.progress_bar.pack(pady=10)
        self.progress_bar.pack_forget()

        self.progress_label = ctk.CTkLabel(btn_frame, text="")
        self.progress_label.pack()
        self.progress_label.pack_forget()

        # Подсказка
        ctk.CTkLabel(frame, text="💾 Настройки сохраняются автоматически", 
                    text_color="gray").pack(pady=(0, 10))

    # ==================== ОБРАБОТЧИКИ ====================

    def _add_row_label(self):
        """Добавляет строку"""
        self.row_labels.append(f"Склад {len(self.row_labels) + 1}")
        self._render_row_labels()
        self._save_config()

    def _remove_row_label(self, index: int):
        """Удаляет строку"""
        if len(self.row_labels) > 1:
            self.row_labels.pop(index)
            self._render_row_labels()
            self._save_config()

    def _update_row_label(self, index: int, value: str):
        """Обновляет название строки"""
        self.row_labels[index] = value
        self._save_config()

    def _render_row_labels(self):
        """Отрисовывает строки"""
        # Очищаем старые виджеты
        for widget in self.row_labels_frame.winfo_children():
            if isinstance(widget, ctk.CTkEntry) or isinstance(widget, ctk.CTkButton):
                if widget.cget("text") != "+ Добавить строку":
                    widget.destroy()

        # Создаём новые
        for i, label in enumerate(self.row_labels):
            frame = ctk.CTkFrame(self.row_labels_frame, fg_color="transparent")
            frame.pack(side="left", padx=2)
            
            entry = ctk.CTkEntry(frame, width=100, placeholder_text="Название")
            entry.insert(0, label)
            entry.pack(side="left")
            entry.bind("<FocusOut>", lambda e, idx=i: self._update_row_label(idx, e.widget.get()))
            
            ctk.CTkButton(frame, text="×", width=30, fg_color="transparent",
                         text_color=("gray10", "#DCE4EE"),
                         command=lambda idx=i: self._remove_row_label(idx)).pack(side="left")

    def _load_files(self):
        """Загружает файлы"""
        files = filedialog.askopenfilenames(
            title="Выберите Excel-файлы",
            filetypes=[("Excel файлы", "*.xlsx *.xls"), ("Все файлы", "*.*")]
        )

        for file_path in files:
            try:
                # Проверяем, не добавлен ли уже
                if any(f.file_path == file_path for f in self.source_files):
                    continue
                
                sf = SourceFile(file_path)
                self.source_files.append(sf)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось загрузить файл {os.path.basename(file_path)}:\n{e}")

        self._render_files()
        self._save_config()

    def _remove_file(self, file_id: str):
        """Удаляет файл"""
        self.source_files = [f for f in self.source_files if f.id != file_id]
        self.column_filters = [f for f in self.column_filters if f.source_file_id != file_id]
        self._render_files()
        self._render_filters()
        self._save_config()

    def _open_file_config(self, file_id: str):
        """Открывает окно настройки файла"""
        file = next((f for f in self.source_files if f.id == file_id), None)
        if not file:
            return

        # Создаём модальное окно
        dialog = FileConfigDialog(self, file)
        self.wait_window(dialog)
        
        self._render_files()
        self._save_config()

    def _update_file_display_name(self, file_id: str, name: str):
        """Обновляет отображаемое имя файла"""
        file = next((f for f in self.source_files if f.id == file_id), None)
        if file:
            file.display_name = name
            self._save_config()

    def _render_files(self):
        """Отрисовывает список файлов"""
        # Очищаем
        for widget in self.files_frame.winfo_children():
            widget.destroy()

        if not self.source_files:
            ctk.CTkLabel(self.files_frame, text="Файлы не загружены", 
                        text_color="gray").pack(pady=10)
            self.add_filter_btn.configure(state="disabled")
            return

        self.add_filter_btn.configure(state="normal")

        for file in self.source_files:
            frame = ctk.CTkFrame(self.files_frame, border_width=1)
            frame.pack(fill="x", pady=5)

            # Имя файла
            inner = ctk.CTkFrame(frame, fg_color="transparent")
            inner.pack(fill="x", padx=10, pady=10)

            ctk.CTkLabel(inner, text=f"📄 {file.file_name}", 
                        font=ctk.CTkFont(weight="bold")).pack(anchor="w")

            # Настройки
            settings = ctk.CTkFrame(inner, fg_color="transparent")
            settings.pack(fill="x", pady=5)

            ctk.CTkLabel(settings, text="Имя для столбцов:").pack(side="left")
            
            name_entry = ctk.CTkEntry(settings, width=150)
            name_entry.insert(0, file.display_name)
            name_entry.pack(side="left", padx=5)
            name_entry.bind("<FocusOut>", lambda e, fid=file.id: 
                           self._update_file_display_name(fid, e.widget.get()))

            ctk.CTkLabel(settings, text="Лист:").pack(side="left", padx=(20, 5))
            
            sheet_menu = ctk.CTkOptionMenu(settings, values=file.available_sheets,
                                          width=120)
            sheet_menu.set(file.sheet_name)
            sheet_menu.pack(side="left")
            sheet_menu.configure(command=lambda v, fid=file.id: self._update_file_sheet(fid, v))

            # Кнопки
            btn_frame = ctk.CTkFrame(inner, fg_color="transparent")
            btn_frame.pack(side="right")

            # Статус настройки
            configured = all([file.mapping.get("vendorColumn"), 
                            file.mapping.get("statusColumn"),
                            file.mapping.get("partNumberColumn")])
            status_text = "✅ Настроен" if configured else "⚠️ Не настроен"
            status_color = "green" if configured else "orange"
            
            ctk.CTkLabel(btn_frame, text=status_text, text_color=status_color).pack(side="left", padx=10)
            
            ctk.CTkButton(btn_frame, text="⚙️ Настроить", width=100,
                         command=lambda fid=file.id: self._open_file_config(fid)).pack(side="left", padx=5)
            
            ctk.CTkButton(btn_frame, text="🗑️", width=40, fg_color="transparent",
                         text_color=("gray10", "#DCE4EE"),
                         command=lambda fid=file.id: self._remove_file(fid)).pack(side="left")

    def _update_file_sheet(self, file_id: str, sheet_name: str):
        """Обновляет лист файла"""
        file = next((f for f in self.source_files if f.id == file_id), None)
        if file:
            file.update_sheet(sheet_name)
            self._render_files()
            self._save_config()

    def _add_column_filter(self):
        """Добавляет фильтр столбца"""
        if not self.source_files:
            return
        
        first_file = self.source_files[0]
        cf = ColumnFilter(first_file.id, first_file.display_name)
        self.column_filters.append(cf)
        self._render_filters()
        self._save_config()

    def _remove_column_filter(self, filter_id: str):
        """Удаляет фильтр"""
        self.column_filters = [f for f in self.column_filters if f.id != filter_id]
        self._render_filters()
        self._save_config()

    def _update_filter(self, filter_id: str, **kwargs):
        """Обновляет фильтр"""
        cf = next((f for f in self.column_filters if f.id == filter_id), None)
        if cf:
            for key, value in kwargs.items():
                setattr(cf, key, value)
            # Автообновление имени столбца
            if "vendor_name" in kwargs or "source_file_id" in kwargs:
                file = next((f for f in self.source_files if f.id == cf.source_file_id), None)
                if file:
                    cf.column_name = f"{cf.vendor_name}_{file.display_name}"
            self._render_filters()
            self._save_config()

    def _render_filters(self):
        """Отрисовывает фильтры"""
        for widget in self.filters_frame.winfo_children():
            widget.destroy()

        if not self.column_filters:
            ctk.CTkLabel(self.filters_frame, text="Добавьте столбцы для сводной таблицы", 
                        text_color="gray").pack(pady=10)
            return

        for i, cf in enumerate(self.column_filters):
            frame = ctk.CTkFrame(self.filters_frame, border_width=1)
            frame.pack(fill="x", pady=5)

            inner = ctk.CTkFrame(frame, fg_color="transparent")
            inner.pack(fill="x", padx=10, pady=10)

            # Заголовок фильтра
            header = ctk.CTkFrame(inner, fg_color="transparent")
            header.pack(fill="x")

            ctk.CTkLabel(header, text=f"Столбец {i+1}", 
                        font=ctk.CTkFont(weight="bold")).pack(side="left")

            name_entry = ctk.CTkEntry(header, width=200, placeholder_text="Название столбца")
            name_entry.insert(0, cf.column_name)
            name_entry.pack(side="left", padx=10)
            name_entry.bind("<FocusOut>", lambda e, fid=cf.id: 
                           self._update_filter(fid, column_name=e.widget.get()))

            ctk.CTkButton(header, text="×", width=30, fg_color="transparent",
                         text_color=("gray10", "#DCE4EE"),
                         command=lambda fid=cf.id: self._remove_column_filter(fid)).pack(side="right")

            # Настройки фильтра
            settings = ctk.CTkFrame(inner, fg_color="transparent")
            settings.pack(fill="x", pady=5)

            # Файл
            ctk.CTkLabel(settings, text="Файл:").grid(row=0, column=0, padx=5, pady=2)
            file_names = [f.display_name for f in self.source_files]
            file_menu = ctk.CTkOptionMenu(settings, values=file_names, width=150)
            file = next((f for f in self.source_files if f.id == cf.source_file_id), None)
            if file:
                file_menu.set(file.display_name)
            file_menu.grid(row=0, column=1, padx=5, pady=2)
            file_menu.configure(command=lambda v, fid=cf.id: self._update_filter(
                fid, source_file_id=next(f.id for f in self.source_files if f.display_name == v)))

            # Вендор
            ctk.CTkLabel(settings, text="Вендор:").grid(row=0, column=2, padx=5, pady=2)
            vendor_entry = ctk.CTkEntry(settings, width=100, placeholder_text="GM")
            vendor_entry.insert(0, cf.vendor_name)
            vendor_entry.grid(row=0, column=3, padx=5, pady=2)
            vendor_entry.bind("<FocusOut>", lambda e, fid=cf.id: 
                             self._update_filter(fid, vendor_name=e.widget.get()))

            # Статус
            ctk.CTkLabel(settings, text="Статус:").grid(row=0, column=4, padx=5, pady=2)
            status_entry = ctk.CTkEntry(settings, width=50, placeholder_text="1")
            status_entry.insert(0, cf.status_value)
            status_entry.grid(row=0, column=5, padx=5, pady=2)
            status_entry.bind("<FocusOut>", lambda e, fid=cf.id: 
                             self._update_filter(fid, status_value=e.widget.get()))

    def _generate_excel(self):
        """Генерирует сводный Excel"""
        # Проверки
        if not self.source_files:
            messagebox.showwarning("Предупреждение", "Загрузите хотя бы один файл")
            return

        if not self.column_filters:
            messagebox.showwarning("Предупреждение", "Добавьте хотя бы один столбец")
            return

        # Проверяем настройку файлов
        unconfigured = [f for f in self.source_files 
                       if not all([f.mapping.get("vendorColumn"), 
                                  f.mapping.get("statusColumn"),
                                  f.mapping.get("partNumberColumn")])]
        if unconfigured:
            messagebox.showwarning("Предупреждение", 
                f"Настройте файлы: {', '.join(f.file_name for f in unconfigured)}")
            return

        # Выбор места сохранения
        output_path = filedialog.asksaveasfilename(
            title="Сохранить сводный Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx")],
            initialfile="summary.xlsx"
        )

        if not output_path:
            return

        # Показываем прогресс
        self.progress_bar.pack()
        self.progress_label.pack()
        self.generate_btn.configure(state="disabled")

        # Запускаем в отдельном потоке
        thread = threading.Thread(target=self._process_and_save, args=(output_path,))
        thread.start()

    def _process_and_save(self, output_path: str):
        """Обрабатывает и сохраняет (в отдельном потоке)"""
        try:
            processed_data = {}

            # Обрабатываем каждый файл
            for i, file in enumerate(self.source_files):
                self.after(0, lambda: self.progress_label.configure(
                    text=f"Обработка файла {i+1}/{len(self.source_files)}: {file.display_name}"))
                self.after(0, lambda: self.progress_bar.set((i + 0.5) / len(self.source_files)))
                
                data = DataProcessor.process_file(file, self.row_labels)
                processed_data[file.id] = data

            # Создаём Excel
            self.after(0, lambda: self.progress_label.configure(text="Создание сводной таблицы..."))
            self.after(0, lambda: self.progress_bar.set(0.9))

            DataProcessor.create_summary_excel(
                self.row_labels, self.column_filters, self.counters,
                processed_data, output_path
            )

            # Успех
            self.after(0, lambda: self.progress_bar.set(1))
            self.after(0, lambda: self.progress_label.configure(text="✅ Файл успешно создан!"))
            self.after(0, lambda: messagebox.showinfo("Готово", f"Файл сохранён:\n{output_path}"))

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Ошибка", f"Не удалось создать файл:\n{e}"))
            self.after(0, lambda: self.progress_label.configure(text=f"❌ Ошибка: {e}"))

        finally:
            self.after(0, lambda: self.generate_btn.configure(state="normal"))

    # ==================== КОНФИГУРАЦИЯ ====================

    def _save_config(self):
        """Сохраняет конфигурацию"""
        config = {
            "row_labels": self.row_labels,
            "source_files": [f.to_dict() for f in self.source_files],
            "column_filters": [f.to_dict() for f in self.column_filters],
            "counters": [c.to_dict() for c in self.counters]
        }

        try:
            config_path = os.path.join(os.path.dirname(__file__), CONFIG_FILE)
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка сохранения конфигурации: {e}")

    def _load_config(self):
        """Загружает конфигурацию"""
        try:
            config_path = os.path.join(os.path.dirname(__file__), CONFIG_FILE)
            if not os.path.exists(config_path):
                return

            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)

            self.row_labels = config.get("row_labels", self.row_labels)

            # Восстанавливаем файлы
            for file_data in config.get("source_files", []):
                sf = SourceFile.from_dict(file_data)
                if sf:
                    self.source_files.append(sf)

            # Восстанавливаем фильтры
            for filter_data in config.get("column_filters", []):
                cf = ColumnFilter.from_dict(filter_data)
                self.column_filters.append(cf)

            # Восстанавливаем счётчики
            for counter_data in config.get("counters", []):
                c = Counter.from_dict(counter_data)
                self.counters.append(c)

            self._render_row_labels()
            self._render_files()
            self._render_filters()

        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")


# ==================== ДИАЛОГ НАСТРОЙКИ ФАЙЛА ====================
class FileConfigDialog(ctk.CTkToplevel):
    """Диалог настройки файла"""

    def __init__(self, parent, file: SourceFile):
        super().__init__(parent)

        self.file = file
        self.result = False

        self.title(f"Настройка: {file.file_name}")
        self.geometry("900x700")
        self.transient(parent)
        self.grab_set()

        self._create_ui()

    def _create_ui(self):
        """Создаёт интерфейс"""
        # Основной контейнер
        main = ctk.CTkScrollableFrame(self)
        main.pack(fill="both", expand=True, padx=10, pady=10)

        # Выбор столбцов
        ctk.CTkLabel(main, text="Выберите столбцы для сопоставления", 
                    font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=10)

        columns_frame = ctk.CTkFrame(main)
        columns_frame.pack(fill="x", pady=5)

        column_letters = [get_column_letter(i) for i in range(1, 27)]

        # Столбцы
        fields = [
            ("Столбец с названием вендора:", "vendorColumn"),
            ("Столбец со статусом (1/0):", "statusColumn"),
            ("Столбец с номером детали:", "partNumberColumn"),
            ("Столбец с данными (опционально):", "dataColumn")
        ]

        self.mapping_vars = {}
        for i, (label, key) in enumerate(fields):
            ctk.CTkLabel(columns_frame, text=label).grid(row=i, column=0, padx=10, pady=5, sticky="e")
            
            var = ctk.StringVar(value=self.file.mapping.get(key, ""))
            self.mapping_vars[key] = var
            
            menu = ctk.CTkOptionMenu(columns_frame, variable=var, width=250,
                                    values=[""] + [
                                        f"{letter} - {self.file.headers[j]}" if j < len(self.file.headers) else letter
                                        for j, letter in enumerate(column_letters)
                                    ])
            menu.grid(row=i, column=1, padx=10, pady=5, sticky="w")

        # Превью
        ctk.CTkLabel(main, text="Превью данных", 
                    font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=(20, 10))

        # Таблица превью
        preview_frame = ctk.CTkFrame(main)
        preview_frame.pack(fill="both", expand=True)

        # Создаём Treeview для превью
        tree_frame = ctk.CTkFrame(preview_frame)
        tree_frame.pack(fill="both", expand=True)

        # Используем tkinter Treeview для таблицы
        style = ttk.Style()
        style.theme_use("clam")
        
        tree = ttk.Treeview(tree_frame, show="headings", height=10)
        
        # Добавляем скроллбар
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Заголовки
        if self.file.headers:
            tree["columns"] = list(range(len(self.file.headers)))
            for i, header in enumerate(self.file.headers):
                tree.heading(i, text=f"{get_column_letter(i+1)}: {header[:20]}")
                tree.column(i, width=100, minwidth=50)

        # Данные
        for row in self.file.preview[1:]:  # Пропускаем заголовок
            tree.insert("", "end", values=row)

        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Кнопки
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkButton(btn_frame, text="Отмена", command=self._cancel, 
                     fg_color="transparent", border_width=1).pack(side="right", padx=5)
        ctk.CTkButton(btn_frame, text="Сохранить", command=self._save).pack(side="right", padx=5)

    def _save(self):
        """Сохраняет настройки"""
        for key, var in self.mapping_vars.items():
            # Извлекаем только букву столбца
            value = var.get()
            if " - " in value:
                value = value.split(" - ")[0]
            self.file.mapping[key] = value

        self.result = True
        self.destroy()

    def _cancel(self):
        """Отменяет изменения"""
        self.destroy()


# ==================== ЗАПУСК ====================
def main():
    app = ExcelMergerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
